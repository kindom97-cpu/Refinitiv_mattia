"""
SCRIPT COMPLETO: Download dataset IPO-ESG
==========================================
Scarica tutte le variabili del modello econometrico per 189 aziende.

VARIABILI:
  Dipendente:  EV/EBITDA (anno IPO+2)
  ESG:         ESG Score totale, E, S, G
  Controlli:   Size (ln Assets), Age, ROE, Revenue Growth, Leverage
  FE:          Sector (Industry FE), IPO Year (Year FE)

INSTALLAZIONE (esegui una volta):
  pip install yfinance pandas openpyxl requests

ESECUZIONE:
  python scarica_tutto.py

TEMPO STIMATO: 15-25 minuti per 189 aziende
"""

import yfinance as yf
import pandas as pd
import numpy as np
import requests
import time
import warnings
warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════
# CONFIGURAZIONE
# ══════════════════════════════════════════════
INPUT_FILE  = "ISIN_IPO.xlsx"
OUTPUT_FILE = "dataset_finale.xlsx"

# ══════════════════════════════════════════════
# STEP 0: CARICA FILE
# ══════════════════════════════════════════════
print("\n" + "═"*60)
print("  DOWNLOAD DATASET IPO-ESG")
print("═"*60)

from openpyxl import load_workbook
wb = load_workbook(INPUT_FILE, data_only=True)
ws = wb.active
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows.append(row)

df = pd.DataFrame(rows, columns=["ISIN","IPO_DATE","LASTDAY_FY1","FY_Y1","FY_Y","c6","c7"])
df = df[["ISIN","IPO_DATE","LASTDAY_FY1","FY_Y1","FY_Y"]]
df["ISIN"]     = df["ISIN"].astype(str).str.strip().str.replace('\xa0','')
df["IPO_DATE"] = pd.to_datetime(df["IPO_DATE"])
df["IPO_YEAR"] = df["IPO_DATE"].dt.year
df["REF_YEAR"] = df["FY_Y1"].str.replace("FY","").astype(int)
df["PREV_YEAR"] = df["FY_Y"].str.replace("FY","").astype(int)

print(f"✅ Caricate {len(df)} aziende")
print(f"   Anni IPO: {sorted(df['IPO_YEAR'].unique())}\n")

# ══════════════════════════════════════════════
# STEP 1: ISIN → TICKER (via OpenFIGI gratuito)
# ══════════════════════════════════════════════
def isin_to_ticker(isin):
    """Converte ISIN in ticker Yahoo Finance usando OpenFIGI API (gratuita)."""
    try:
        url = "https://api.openfigi.com/v3/mapping"
        headers = {"Content-Type": "application/json"}
        payload = [{"idType": "ID_ISIN", "idValue": isin}]
        r = requests.post(url, json=payload, headers=headers, timeout=10)
        if r.status_code == 200:
            data = r.json()
            if data and data[0].get("data"):
                # Preferisci exchange US
                for item in data[0]["data"]:
                    if item.get("exchCode") in ["US", "UN", "UQ", "UA", "UW"]:
                        return item.get("ticker")
                return data[0]["data"][0].get("ticker")
    except Exception:
        pass
    return None

print("STEP 1: Conversione ISIN → Ticker...")
tickers = []
for i, row in df.iterrows():
    ticker = isin_to_ticker(row["ISIN"])
    tickers.append(ticker)
    stato = f"✅ {ticker}" if ticker else "❌ non trovato"
    print(f"  [{i+1:3d}/189] {row['ISIN']} → {stato}")
    time.sleep(0.4)  # rispetta rate limit OpenFIGI (25 req/min free tier)

df["TICKER"] = tickers
n_ok = df["TICKER"].notna().sum()
print(f"\n  Trovati: {n_ok}/189 ticker\n")

# ══════════════════════════════════════════════
# STEP 2: DOWNLOAD DATI FINANZIARI
# ══════════════════════════════════════════════
def get_all_data(ticker, ref_year, prev_year, ipo_year):
    """
    Scarica tutte le variabili per un ticker dato.
    ref_year  = anno IPO + 2  (es. 2022 per Airbnb)
    prev_year = anno IPO + 1  (es. 2021 per Airbnb, per calcolare growth)
    """
    out = {
        "ev_ebitda": None,
        "esg_score": None, "esg_env": None, "esg_soc": None, "esg_gov": None,
        "total_assets": None, "size_log": None,
        "roe": None,
        "rev_y": None, "rev_y1": None, "revenue_growth": None,
        "leverage": None,
        "sector": None, "industry": None,
        "founded_year": None, "age": None,
        "note": ""
    }

    try:
        stk = yf.Ticker(ticker)
        info = stk.info or {}

        # ── SECTOR & INDUSTRY ─────────────────
        out["sector"]   = info.get("sector")
        out["industry"] = info.get("industry")

        # ── ESG SCORE ─────────────────────────
        try:
            esg = stk.sustainability
            if esg is not None and not esg.empty:
                def esg_val(key):
                    if key in esg.index:
                        v = esg.loc[key].iloc[0]
                        return float(v) if pd.notna(v) else None
                    return None
                out["esg_score"] = esg_val("totalEsg")
                out["esg_env"]   = esg_val("environmentScore")
                out["esg_soc"]   = esg_val("socialScore")
                out["esg_gov"]   = esg_val("governanceScore")
        except Exception:
            pass

        # ── DATI FINANZIARI STORICI ───────────
        def find_col(df_fin, year):
            """Trova la colonna corrispondente all'anno nel dataframe finanziario."""
            if df_fin is None or df_fin.empty:
                return None
            for col in df_fin.columns:
                try:
                    if pd.Timestamp(col).year == year:
                        return col
                except Exception:
                    pass
            return None

        def safe_val(df_fin, fields, col):
            """Legge il valore da un dataframe finanziario in modo sicuro."""
            if df_fin is None or col is None:
                return None
            for f in fields:
                if f in df_fin.index:
                    v = df_fin.loc[f, col]
                    if pd.notna(v):
                        return float(v)
            return None

        inc  = stk.financials
        bal  = stk.balance_sheet

        col_ref  = find_col(inc, ref_year)
        col_prev = find_col(inc, prev_year)
        col_bal  = find_col(bal, ref_year)

        # Revenue anno Y e Y-1
        rev_fields = ["Total Revenue", "Revenue", "Operating Revenue"]
        rev_y  = safe_val(inc, rev_fields, col_ref)
        rev_y1 = safe_val(inc, rev_fields, col_prev)
        if rev_y:  out["rev_y"]  = rev_y  / 1e6
        if rev_y1: out["rev_y1"] = rev_y1 / 1e6

        # Growth
        if rev_y and rev_y1 and rev_y1 != 0:
            out["revenue_growth"] = (rev_y - rev_y1) / abs(rev_y1) * 100

        # EBITDA
        ebitda_fields = ["EBITDA", "Normalized EBITDA", "Reconciled Depreciation"]
        ebitda = safe_val(inc, ebitda_fields, col_ref)
        ebitda_m = ebitda / 1e6 if ebitda else None

        # Total Assets
        asset_fields = ["Total Assets"]
        assets = safe_val(bal, asset_fields, col_bal)
        if assets and assets > 0:
            out["total_assets"] = assets / 1e6
            out["size_log"]     = np.log(assets / 1e6)

        # Equity e Debt per ROE e Leverage
        equity_fields = ["Stockholders Equity", "Total Stockholder Equity",
                         "Common Stock Equity", "Total Equity Gross Minority Interest"]
        debt_fields   = ["Total Debt", "Long Term Debt And Capital Lease Obligation"]
        equity = safe_val(bal, equity_fields, col_bal)
        debt   = safe_val(bal, debt_fields,   col_bal)

        # ROE
        ni_fields = ["Net Income", "Net Income Common Stockholders",
                     "Net Income Including Noncontrolling Interests"]
        net_income = safe_val(inc, ni_fields, col_ref)
        if net_income and equity and equity != 0:
            out["roe"] = (net_income / equity) * 100

        # Leverage
        if debt is not None and equity and equity != 0:
            out["leverage"] = debt / equity

        # EV/EBITDA storico (calcolato)
        ev_info = info.get("enterpriseValue")
        if ev_info and ebitda_m and ebitda_m != 0:
            out["ev_ebitda"] = (ev_info / 1e6) / ebitda_m
        elif info.get("enterpriseToEbitda"):
            out["ev_ebitda"] = info.get("enterpriseToEbitda")

        # Founded Year & Age
        founded = info.get("foundedYear") or info.get("fundInceptionDate")
        if founded:
            try:
                fy = int(str(founded)[:4])
                out["founded_year"] = fy
                out["age"] = ipo_year - fy
            except Exception:
                pass

    except Exception as e:
        out["note"] = str(e)[:80]

    return out

print("STEP 2: Download dati finanziari (yfinance)...")
all_results = []

for i, row in df.iterrows():
    ticker    = row["TICKER"]
    ref_year  = int(row["REF_YEAR"])
    prev_year = int(row["PREV_YEAR"])
    ipo_year  = int(row["IPO_YEAR"])

    if not ticker:
        all_results.append({})
        print(f"  [{i+1:3d}/189] saltato — nessun ticker")
        continue

    data = get_all_data(ticker, ref_year, prev_year, ipo_year)
    all_results.append(data)

    ev  = f"{data['ev_ebitda']:.1f}x" if data['ev_ebitda'] else "N/D"
    esg = f"{data['esg_score']:.1f}"  if data['esg_score'] else "N/D"
    roe = f"{data['roe']:.1f}%"       if data['roe']       else "N/D"
    grw = f"{data['revenue_growth']:.1f}%" if data['revenue_growth'] else "N/D"
    print(f"  [{i+1:3d}/189] {ticker:8s} FY{ref_year} → EV/EBITDA:{ev:8s} ESG:{esg:6s} ROE:{roe:8s} Growth:{grw}")
    time.sleep(0.2)

# ══════════════════════════════════════════════
# STEP 3: ASSEMBLA DATASET FINALE
# ══════════════════════════════════════════════
print("\nSTEP 3: Assemblo il dataset finale...")

df_res = pd.DataFrame(all_results)
dataset = pd.concat([df.reset_index(drop=True), df_res], axis=1)

# Colonne finali ordinate
cols = [
    "ISIN", "TICKER", "IPO_DATE", "IPO_YEAR", "REF_YEAR",
    "ev_ebitda",
    "esg_score", "esg_env", "esg_soc", "esg_gov",
    "size_log", "total_assets",
    "age", "founded_year",
    "roe",
    "revenue_growth", "rev_y", "rev_y1",
    "leverage",
    "sector", "industry",
    "note"
]
cols_ok = [c for c in cols if c in dataset.columns]
final = dataset[cols_ok].copy()

# ══════════════════════════════════════════════
# STEP 4: SALVA EXCEL CON PIÙ FOGLI
# ══════════════════════════════════════════════
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    # Foglio 1: Dataset completo
    final.to_excel(writer, sheet_name="Dataset", index=False)

    # Foglio 2: Solo osservazioni complete per la regressione
    vars_reg = ["ev_ebitda","esg_score","size_log","age","roe","revenue_growth","leverage"]
    vars_ok  = [v for v in vars_reg if v in final.columns]
    completo = final.dropna(subset=vars_ok)
    completo.to_excel(writer, sheet_name="Regressione", index=False)

    # Foglio 3: Statistiche descrittive
    final[vars_ok].describe().round(3).to_excel(writer, sheet_name="Statistiche")

    # Foglio 4: Missing values
    miss = final.isnull().sum().reset_index()
    miss.columns = ["Variabile", "N_Missing"]
    miss["Pct_Missing"] = (miss["N_Missing"] / len(final) * 100).round(1)
    miss.to_excel(writer, sheet_name="Missing", index=False)

# ══════════════════════════════════════════════
# RIEPILOGO FINALE
# ══════════════════════════════════════════════
print("\n" + "═"*60)
print("  COMPLETATO!")
print("═"*60)
print(f"  Totale aziende:          {len(final)}")
print(f"  Ticker trovati:          {final['TICKER'].notna().sum()}")
print(f"  EV/EBITDA disponibili:   {final['ev_ebitda'].notna().sum()}")
print(f"  ESG Score disponibili:   {final['esg_score'].notna().sum()}")
print(f"  Osservazioni complete:   {len(completo)} (foglio 'Regressione')")
print(f"\n  File salvato: {OUTPUT_FILE}")
print("═"*60)
print("""
⚠️  NOTA SU EV/EBITDA:
   yfinance non ha EV/EBITDA storico per l'anno IPO+2.
   Il valore scaricato è il più recente disponibile.
   Se hai ancora accesso parziale a Refinitiv anche solo
   per 10 minuti, usa il file definitivo che abbiamo
   costruito insieme per scaricare solo EV/EBITDA.
""")
